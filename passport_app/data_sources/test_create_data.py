import os
import sys
import django
import requests

sys.path.append("D:\Work Projects\passport_app\property_passport")
os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'property_passport.settings')
django.setup()

from passport_app.models import *

from openpyxl import load_workbook

def create_parser_param(name, name_ru, parser_type_id, parser_parameter_type):
    pp = ParserParameter()
    pp.name = name
    pp.name_ru = name_ru
    pp.parser_type_id = parser_type_id
    pp.parser_parameter_type = parser_parameter_type

    pp.save()
    return pp

def create_param(name, name_ru, parser_parameters):
    p = Parameter()
    p.name = name
    p.name_ru = name_ru

    p.save()

    for param in parser_parameters:
        p.parser_parameters.add(param)

    p.save()
    return p


def create_category(wsheet, row):
    index = str(wsheet.cell(row, 1).value)
    next_index = str(wsheet.cell(row+1, 1).value)
    if next_index == 'None':
        next_index = index

    param_name_rus = wsheet.cell(row, 4).value
    param_name = wsheet.cell(row, 5).value

    parser_param_name_rus = wsheet.cell(row, 6).value
    parser_param_name = wsheet.cell(row, 7).value

    #add category
    category = Category()
    category.name_ru = wsheet.cell(row, 2).value
    category.name = wsheet.cell(row, 3).value
    category.point = index
    category.comment = index.replace(".", "")
    category.save()

    if len(index.split('.')) > 1:
        parent_index = '.'.join(index.split('.')[:-1])
        parent = Category.objects.filter(point = parent_index).first()

        if parent_index == '1':
            parent = Category.objects.first()
            parent.categories.add(category)
            parent.save()
        else:
            parent.categories.add(category)

        category.parent_categories.add(parent)
    category.save()

    #add param to category
    if parser_param_name:
        if len(index.split('.')) >= len(next_index.split(';')[0].split('.')) and len(index.split('.')) != 1:
            param = Parameter.objects.filter(name = param_name).first()
            if param is not None:
                print('ERROR')
                category.parameters.add(param)
            else:
                print("param '" + category.name + "' not found")

                parser_params = ParserParameter.objects.filter(name = parser_param_name).all()
                if len(parser_params) > 1:
                    print('ERROR')

                if len(parser_params) == 0:
                    print('parser_param not found')
                    parser_params = []
                    if wsheet.cell(row, 8).value == '+':
                        p = create_parser_param(parser_param_name + " Google", parser_param_name_rus + " Google", 2, 'social')
                        parser_params.append(p)

                        p = create_parser_param(parser_param_name + " Yandex", parser_param_name_rus + " Yandex", 3, 'social')
                        parser_params.append(p)
                    elif wsheet.cell(row, 10).value == '+': 
                        p = create_parser_param(parser_param_name + " Avito", parser_param_name_rus + " Avito", 26, 'social')
                        parser_params.append(p)

                        p = create_parser_param(parser_param_name + " Cian", parser_param_name_rus + " Cian", 27, 'social')
                        parser_params.append(p)
                    else:
                        parser_params = [create_parser_param(parser_param_name, parser_param_name_rus, wsheet.cell(row, 9).value, '')]

                param = create_param(param_name, param_name_rus, parser_params)
                category.parameters.add(param)            

            category.save()
            print('addedd')

def start(file_name):
    wb = None
    wsheet = None
    
    wb = load_workbook(file_name)
    wsheet = wb.get_sheet_by_name(wb.get_sheet_names()[0])

    max_row = 1
    while wsheet.cell(max_row, 1).value is not None:
        max_row = max_row + 1

    for row in range(2, max_row):
        print(row)
        create_category(wsheet, row)

    wb.close()
        
def delete_all():
    cat = Category.objects.all().delete()
    Parameter.objects.all().delete()
    ParserParameter.objects.all().delete()

def create_formulas(category: Category, parent_category: Category):
    if category.categories.count() != 0:
        for cat in category.categories.all():
            create_formulas(cat, category)

        formula = FormulaCategory()
        formula.category = category

        arr = category.point.split('.')
        if len(arr) == 2:
            formula.formula = 'avrg(%s-%s)' % (category.categories.first().point, category.categories.last().point)
        else:
            formula.formula = 'sum(%s-%s)' % (category.categories.first().point, category.categories.last().point)
        
        formula.save()
    else:
        formula = FormulaCategory()
        formula.category = category
        formula.rate = 5
        formula.amount = parent_category.categories.count()
        formula.formula = 'x/y'

        formula.save()

def create_rate_classifier(wsheet, row):
    category = Category.objects.filter(name_ru__contains = wsheet.cell(row, 1).value).first()
    
    obj = RateClassifier()
    obj.category = category
    obj.label = wsheet.cell(row, 2).value
    obj.min_rate = wsheet.cell(row, 3).value
    obj.max_rate = wsheet.cell(row, 4).value

    obj.save()

def start_create_rate_classifier(file_name):
    wb = load_workbook(file_name)
    wsheet = wb.get_sheet_by_name(wb.get_sheet_names()[0])

    max_row = 1
    while wsheet.cell(max_row, 1).value is not None:
        max_row = max_row + 1

    for row in range(2, max_row):
        print(row)
        create_rate_classifier(wsheet, row)



delete_all()
start(r"C:\Users\Dmitriev Ivan\Desktop\парсинг питон\набор 1, категории для загрузки.xlsx")
start(r"C:\Users\Dmitriev Ivan\Desktop\парсинг питон\набор 2, категории для загрузки.xlsx")
start(r"C:\Users\Dmitriev Ivan\Desktop\парсинг питон\набор 3, категории для загрузки.xlsx")

start_create_rate_classifier(r"C:\Users\Dmitriev Ivan\Desktop\парсинг питон\classifier.xlsx")

categories = Category.objects.all()
search_form = SearchForm.objects.get(name= 'default')

for c in categories:
    search_form.categories.add(c)

search_form.save()

FormulaCategory.objects.all().delete()
categories = Category.objects.filter(parent_categories = None)
for c in categories:
    create_formulas(c, None)
